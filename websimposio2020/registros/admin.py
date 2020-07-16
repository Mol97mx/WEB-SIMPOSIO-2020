from django.contrib import admin
from django.http import HttpResponse
from .models import registro
# Register your models here.

def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=registros.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"Nombre o Razón social"),
        smart_str(u"RFC"),
        smart_str(u"Calle"),
        smart_str(u"Numero exterior"),
        smart_str(u"Numero interior"),
        smart_str(u"Colonia"),
        smart_str(u"Municipio"),
        smart_str(u"Estado"),
        smart_str(u"Pais"),
        smart_str(u"Email"),
        smart_str(u"Telefono"),
        smart_str(u"Uso de CFDI"),
        smart_str(u"Forma de pago"),
        smart_str(u"Método de pago"),
        smart_str(u"Opción"),
        smart_str(u"Estatus de pago"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.nombre),
            smart_str(obj.rfc),
            smart_str(obj.calle),
            smart_str(obj.numeroexterior),
            smart_str(obj.numerointerior),
            smart_str(obj.colonia),
            smart_str(obj.ciudad),
            smart_str(obj.estado),
            smart_str(obj.pais),
            smart_str(obj.email),
            smart_str(obj.telefono),
            smart_str(obj.cfdi),
            smart_str(obj.formadepago),
            smart_str(obj.metododepago),
            smart_str(obj.opcion),
            smart_str(obj.pagado),
        ])
    return response
export_csv.short_description = u"Export CSV"

def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=registros.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("Registros")
    
    row_num = 0
    
    columns = [
        (u"Nombre o Razón social",8000),
        (u"RFC",6000),
        (u"Calle",3000),
        (u"Numero exterior",3000),
        (u"Numero interior",3000),
        (u"Colonia",6000),
        (u"Municipio",6000),
        (u"Estado",6000),
        (u"Pais",3000),
        (u"Email",6000),
        (u"Telefono",6000),
        (u"Uso de CFDI",6000),
        (u"Forma de pago",6000),
        (u"Método de pago",6000),
        (u"Opción",8000),
        (u"Estatus de pago",6000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = [
            obj.nombre,
            obj.rfc,
            obj.calle,
            obj.numeroexterior,
            obj.numerointerior,
            obj.colonia,
            obj.ciudad,
            obj.estado,
            obj.pais,
            obj.email,
            obj.telefono,
            obj.cfdi,
            obj.formadepago,
            obj.metododepago,
            obj.opcion,
            obj.pagado,
        ]
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response
    
export_xls.short_description = u"Export XLS"

def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registros.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    row_num = 0

    columns = [
        (u"Nombre o Razón social",80),
        (u"RFC",60),
        (u"Calle",30),
        (u"Numero exterior",30),
        (u"Numero interior",30),
        (u"Colonia",60),
        (u"Municipio",60),
        (u"Estado",60),
        (u"Pais",30),
        (u"Email",60),
        (u"Telefono",60),
        (u"Uso de CFDI",60),
        (u"Forma de pago",60),
        (u"Método de pago",60),
        (u"Opción",80),
        (u"Estatus de pago",60),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.nombre,
            obj.rfc,
            obj.calle,
            obj.numeroexterior,
            obj.numerointerior,
            obj.colonia,
            obj.ciudad,
            obj.estado,
            obj.pais,
            obj.email,
            obj.telefono,
            obj.cfdi,
            obj.formadepago,
            obj.metododepago,
            obj.opcion,
            obj.pagado,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            # c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"

class registroAdmin(admin.ModelAdmin):
    readonly_fields=('fechaderegistro','fechademodificacion')
    list_display = ('nombre','email','opcion','pagado')
    ordering = ('nombre','rfc','ciudad','estado')
    search_fields=('nombre','email')
    list_filter=('nombre','email','opcion')
    actions = [export_csv, export_xls, export_xlsx]



admin.site.register(registro,registroAdmin)